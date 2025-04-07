import openpyxl
from models.car import CAuto
import sys


class CGarage:
    def __init__(self, max_auto):
        self.cautos = []
        self.max_auto = max_auto

    def add_car(self, car, from_file=False):
        if from_file or len(self.cautos) < self.max_auto:
            self.cautos.append(car)
            return True
        return False

    def save_to_file(self, filename):
        try:
            try:
                wb = openpyxl.load_workbook(filename)
                ws = wb.active
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Car Garage Data"
                headers = [
                    "License Plate",
                    "Model",
                    "Color",
                    "Owner",
                ]
                ws.append(headers)

            for car in self.cautos:
                ws.append(
                    [
                        car.license_plate,
                        car.model,
                        car.color,
                        car.owner,
                    ]
                )

            wb.save(filename)
            return True
        except Exception as e:
            print(f"Ошибка при сохранении в файл: {e}")
            return False

    def load_from_file(self, filename):
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 5 and all(cell is not None for cell in row):
                    id, license_plate, model, color, owner = row
                    car = CAuto(license_plate, model, color, owner)
                    car.id = id
                    self.add_car(car, from_file=True)  # Передаем from_file=True
                    print(f"Загружен автомобиль: {car}")
                else:
                    print(f"Пропущена строка: {row}")
        except FileNotFoundError:
            print(f"Файла {filename} не существует!")
            sys.exit()

    def sort_by_owner(self):
        self.cautos.sort(key=lambda car: car.owner)
